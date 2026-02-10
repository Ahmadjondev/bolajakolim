"""Excel file generation service for Xarajatlar Smetasi.

Mirrors the frontend excelGenerator.ts logic, producing
the same multi-sheet workbook server-side.
"""

import logging
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from django.conf import settings
from django.core.files.base import ContentFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from smetalar.models import XarajatlarSmetasi

logger = logging.getLogger(__name__)

SOCIAL_TAX_RATE = Decimal("0.12")
PROFIT_TAX_RATE = Decimal("0.12")

# ------------------------------------------------------------------ Styles
_THIN = Side(style="thin")
_BORDER = Border(top=_THIN, left=_THIN, bottom=_THIN, right=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="4A86E8")
_HEADER_FONT = Font(bold=True, size=11)
_TITLE_FONT = Font(bold=True, size=14)
_BOLD_FONT = Font(bold=True)
_ITALIC_FONT = Font(italic=True)
_CENTER = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)
_LEFT = Alignment(
    horizontal="left",
    vertical="center",
    wrap_text=True,
)
_RIGHT = Alignment(
    horizontal="right",
    vertical="center",
    wrap_text=True,
)


def _header_style(ws: Any, row: int, col: int, value: str) -> None:
    """Apply header style to a cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = _CENTER
    cell.border = _BORDER


def _cell(
    ws: Any,
    row: int,
    col: int,
    value: Any,
    bold: bool = False,
    align: str = "center",
) -> None:
    """Write a bordered cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = _BORDER
    cell.font = _BOLD_FONT if bold else Font()
    if align == "left":
        cell.alignment = _LEFT
    elif align == "right":
        cell.alignment = _RIGHT
    else:
        cell.alignment = _CENTER


def _rnd(val: float) -> int:
    """Round to nearest integer (for ming so'mda display)."""
    return round(val)


# ------------------------------------------------------------------
# Data gathering helpers
# ------------------------------------------------------------------
def _gather_smeta_data(smeta: XarajatlarSmetasi) -> dict[str, Any]:
    """Prefetch and compute all values needed for the workbook.

    Args:
        smeta: The XarajatlarSmetasi instance.

    Returns:
        Dict with all computed totals and raw data lists.
    """
    employees = list(smeta.employees.all())
    inventory = list(smeta.inventory_items.all())
    raw_materials = list(smeta.raw_materials.all())
    other_expenses = list(smeta.other_expenses.all())
    products = list(smeta.products.all())
    davr = list(smeta.davr_xarajatlari.all())
    sotish_yillari = list(
        smeta.sotish_rejasi_yillari.prefetch_related("products").all()
    )

    mgmt_staff = [e for e in employees if e.staff_type == "management"]
    prod_staff = [e for e in employees if e.staff_type == "production"]

    def _emp_total(emp: Any) -> float:
        return float(emp.monthly_salary) * emp.count * emp.duration_months

    mgmt_total = sum(_emp_total(e) for e in mgmt_staff)
    prod_total = sum(_emp_total(e) for e in prod_staff)

    mgmt_vaz = sum(
        _emp_total(e) for e in mgmt_staff if e.financing_source == "vazirlik"
    )
    mgmt_tash = sum(
        _emp_total(e) for e in mgmt_staff if e.financing_source == "tashkilot"
    )
    prod_vaz = sum(
        _emp_total(e) for e in prod_staff if e.financing_source == "vazirlik"
    )
    prod_tash = sum(
        _emp_total(e) for e in prod_staff if e.financing_source == "tashkilot"
    )

    total_salary = mgmt_total + prod_total
    total_social = total_salary * float(SOCIAL_TAX_RATE)

    inv_total = sum(float(i.price) * i.quantity for i in inventory)
    inv_vaz = sum(
        float(i.price) * i.quantity
        for i in inventory
        if i.financing_source == "vazirlik"
    )
    inv_tash = sum(
        float(i.price) * i.quantity
        for i in inventory
        if i.financing_source == "tashkilot"
    )

    rm_total = sum(float(r.price) * r.quantity for r in raw_materials)
    rm_vaz = sum(
        float(r.price) * r.quantity
        for r in raw_materials
        if r.financing_source == "vazirlik"
    )
    rm_tash = sum(
        float(r.price) * r.quantity
        for r in raw_materials
        if r.financing_source == "tashkilot"
    )

    mgmt_exp = [o for o in other_expenses if o.expense_type == "management"]
    prod_exp = [o for o in other_expenses if o.expense_type == "production"]
    oe_total = sum(float(o.price) * o.quantity for o in other_expenses)
    oe_mgmt_total = sum(float(o.price) * o.quantity for o in mgmt_exp)
    oe_prod_total = sum(float(o.price) * o.quantity for o in prod_exp)
    oe_mgmt_vaz = sum(
        float(o.price) * o.quantity
        for o in mgmt_exp
        if o.financing_source == "vazirlik"
    )
    oe_mgmt_tash = sum(
        float(o.price) * o.quantity
        for o in mgmt_exp
        if o.financing_source == "tashkilot"
    )
    oe_prod_vaz = sum(
        float(o.price) * o.quantity
        for o in prod_exp
        if o.financing_source == "vazirlik"
    )
    oe_prod_tash = sum(
        float(o.price) * o.quantity
        for o in prod_exp
        if o.financing_source == "tashkilot"
    )

    davr_total = sum(float(d.amount) for d in davr)

    grand_total = total_salary + total_social + inv_total + rm_total + oe_total

    # Tannarx
    amortizatsiya = inv_total * 0.2
    jami_tannarx = (
        prod_total
        + prod_total * float(SOCIAL_TAX_RATE)
        + rm_total
        + amortizatsiya
        + oe_prod_total
    )
    total_products_qty = sum(p.quantity for p in products)

    project_years = smeta.project_duration_years or 1

    return {
        "smeta": smeta,
        "employees": employees,
        "mgmt_staff": mgmt_staff,
        "prod_staff": prod_staff,
        "inventory": inventory,
        "raw_materials": raw_materials,
        "other_expenses": other_expenses,
        "mgmt_exp": mgmt_exp,
        "prod_exp": prod_exp,
        "products": products,
        "davr": davr,
        "sotish_yillari": sotish_yillari,
        "mgmt_total": mgmt_total,
        "prod_total": prod_total,
        "mgmt_vaz": mgmt_vaz,
        "mgmt_tash": mgmt_tash,
        "prod_vaz": prod_vaz,
        "prod_tash": prod_tash,
        "total_salary": total_salary,
        "total_social": total_social,
        "inv_total": inv_total,
        "inv_vaz": inv_vaz,
        "inv_tash": inv_tash,
        "rm_total": rm_total,
        "rm_vaz": rm_vaz,
        "rm_tash": rm_tash,
        "oe_total": oe_total,
        "oe_mgmt_total": oe_mgmt_total,
        "oe_prod_total": oe_prod_total,
        "oe_mgmt_vaz": oe_mgmt_vaz,
        "oe_mgmt_tash": oe_mgmt_tash,
        "oe_prod_vaz": oe_prod_vaz,
        "oe_prod_tash": oe_prod_tash,
        "davr_total": davr_total,
        "grand_total": grand_total,
        "amortizatsiya": amortizatsiya,
        "jami_tannarx": jami_tannarx,
        "total_products_qty": total_products_qty,
        "project_years": project_years,
    }


# ------------------------------------------------------------------
# Sheet builders
# ------------------------------------------------------------------
def _build_jami_sheet(wb: Workbook, d: dict) -> None:
    """Build the 'Jami' (summary) sheet."""
    ws = wb.active
    ws.title = "Jami"

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = (
        "Loyihaning xarajatlar smetasi (xarajatlar smetasi loyihaning "
        "umumiy muddatiga to'ldirilishi va xarajatlarni asoslovchi "
        "hisob-kitoblar (jadvallar) ilova qilinishi shart)."
    )
    c.font = Font(size=10, italic=True)
    c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 50

    ws.merge_cells("A4:E4")
    ws["A4"].value = "XARAJATLAR SMETASI"
    ws["A4"].font = _TITLE_FONT
    ws["A4"].alignment = _CENTER

    ws.merge_cells("A5:E5")
    name = d["smeta"].project_name
    org = d["smeta"].organization_name
    ws["A5"].value = f"{name} - {org}"
    ws["A5"].font = Font(bold=True, size=12)
    ws["A5"].alignment = _CENTER

    headers = [
        "Xarajat turlari",
        "Vazirlik hisobidan\n(ming so'mda)",
        "Birgalikda moliyalashtiradigan\ntashkilot hisobidan\n(ming so'mda)",
        "Summa\n(ming so'mda)",
        "Umumiy\nxarajatlardagi\nulushi foizda (%)",
    ]
    for i, h in enumerate(headers, 1):
        _header_style(ws, 7, i, h)
    ws.row_dimensions[7].height = 45

    vaz_salary = d["mgmt_vaz"] + d["prod_vaz"]
    tash_salary = d["mgmt_tash"] + d["prod_tash"]
    vaz_social = vaz_salary * float(SOCIAL_TAX_RATE)
    tash_social = tash_salary * float(SOCIAL_TAX_RATE)

    expenses = [
        ("Ish haqi fondi", vaz_salary, tash_salary),
        ("Ijtimoiy soliq", vaz_social, tash_social),
        (
            "Xomashyo va materiallarni sotib olish bilan bog'liq xarajatlar",
            d["rm_vaz"],
            d["rm_tash"],
        ),
        (
            "Asbob-uskuna, texnika va jihozlarni xarid qilish xarajatlari",
            d["inv_vaz"],
            d["inv_tash"],
        ),
        (
            "Boshqa xarajatlar",
            d["oe_mgmt_vaz"] + d["oe_prod_vaz"],
            d["oe_mgmt_tash"] + d["oe_prod_tash"],
        ),
    ]

    row = 8
    gt = d["grand_total"]
    t_vaz = 0.0
    t_tash = 0.0
    for name, vaz, tash in expenses:
        s = vaz + tash
        pct = f"{(s / gt * 100):.1f}" if gt else "0.0"
        _cell(ws, row, 1, name, align="left")
        _cell(ws, row, 2, _rnd(vaz / 1000))
        _cell(ws, row, 3, _rnd(tash / 1000))
        _cell(ws, row, 4, _rnd(s / 1000))
        _cell(ws, row, 5, f"{pct}%")
        t_vaz += vaz
        t_tash += tash
        row += 1

    _cell(ws, row, 1, "Jami xarajatlar:", bold=True, align="right")
    _cell(ws, row, 2, _rnd(t_vaz / 1000), bold=True)
    _cell(ws, row, 3, _rnd(t_tash / 1000), bold=True)
    _cell(ws, row, 4, _rnd((t_vaz + t_tash) / 1000), bold=True)
    _cell(ws, row, 5, "100.0%", bold=True)

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18


def _build_ish_haqi_sheet(wb: Workbook, d: dict) -> None:
    """Build the 'Ish haqi' (salary) sheet."""
    ws = wb.create_sheet("Ish haqi")

    ws.merge_cells("A1:I1")
    ws["A1"].value = "Mehnatga haq to'lash xarajatlari"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].fill = _HEADER_FILL
    ws["A1"].alignment = _CENTER

    headers = [
        "N",
        "Lavozimi",
        "Ishchilar soni",
        "Bir ishchining\nish haqi\n(oyda)",
        "Jami oylik ish\nhaqi\n(ming so'mda)",
        "Ish davomiyligi\n(oylarda)",
        "Jami\n(ming so'mda)",
    ]
    for i, h in enumerate(headers, 1):
        ws.merge_cells(start_row=3, start_column=i, end_row=4, end_column=i)
        _header_style(ws, 3, i, h)
    ws.merge_cells("H3:I3")
    _header_style(ws, 3, 8, "Moliyalashtirish manbasi")
    _header_style(ws, 4, 8, "Vazirlik hisobidan\n(ming so'mda)")
    _header_style(ws, 4, 9, "Tashkilot hisobidan\n(ming so'mda)")

    row = 5

    def _write_staff(staff: list, label: str, start_row: int) -> int:
        ws.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=9,
        )
        ws.cell(start_row, 1, label).font = _BOLD_FONT
        r = start_row + 1
        for idx, emp in enumerate(staff, 1):
            t = float(emp.monthly_salary) * emp.count * emp.duration_months
            mt = float(emp.monthly_salary) * emp.count
            is_v = emp.financing_source == "vazirlik"
            _cell(ws, r, 1, idx)
            _cell(ws, r, 2, emp.position, align="left")
            _cell(ws, r, 3, emp.count)
            _cell(ws, r, 4, _rnd(float(emp.monthly_salary) / 1000))
            _cell(ws, r, 5, _rnd(mt / 1000))
            _cell(ws, r, 6, emp.duration_months)
            _cell(ws, r, 7, _rnd(t / 1000))
            _cell(ws, r, 8, _rnd(t / 1000) if is_v else "")
            _cell(ws, r, 9, _rnd(t / 1000) if not is_v else "")
            r += 1
        return r

    row = _write_staff(
        d["mgmt_staff"],
        "Ma'muriy-boshqaruv xodimlari:",
        row,
    )
    # Mgmt subtotal
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row, 1, "Jami ma'muriy-boshqaruv xodimlari ish haqi fondi")
    ws.cell(row, 1).font = Font(bold=True, italic=True)
    _cell(ws, row, 7, _rnd(d["mgmt_total"] / 1000), bold=True)
    _cell(ws, row, 8, _rnd(d["mgmt_vaz"] / 1000), bold=True)
    _cell(ws, row, 9, _rnd(d["mgmt_tash"] / 1000), bold=True)
    row += 1

    # Mgmt social tax
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row, 1, "Ijtimoiy soliq").font = _ITALIC_FONT
    st = float(SOCIAL_TAX_RATE)
    _cell(ws, row, 7, _rnd(d["mgmt_total"] * st / 1000))
    _cell(ws, row, 8, _rnd(d["mgmt_vaz"] * st / 1000))
    _cell(ws, row, 9, _rnd(d["mgmt_tash"] * st / 1000))
    row += 1

    row = _write_staff(
        d["prod_staff"],
        "Ishlab chiqarish xodimlari:",
        row,
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row, 1, "Jami ishlab chiqarish xodimlari ish haqi fondi")
    ws.cell(row, 1).font = Font(bold=True, italic=True)
    _cell(ws, row, 7, _rnd(d["prod_total"] / 1000), bold=True)
    _cell(ws, row, 8, _rnd(d["prod_vaz"] / 1000), bold=True)
    _cell(ws, row, 9, _rnd(d["prod_tash"] / 1000), bold=True)
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row, 1, "Ijtimoiy soliq").font = _ITALIC_FONT
    _cell(ws, row, 7, _rnd(d["prod_total"] * st / 1000))
    _cell(ws, row, 8, _rnd(d["prod_vaz"] * st / 1000))
    _cell(ws, row, 9, _rnd(d["prod_tash"] * st / 1000))
    row += 1

    # Grand totals
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row, 1, "Jami ish haqi fondi").font = _BOLD_FONT
    _cell(ws, row, 7, _rnd(d["total_salary"] / 1000), bold=True)
    vaz_all = d["mgmt_vaz"] + d["prod_vaz"]
    tash_all = d["mgmt_tash"] + d["prod_tash"]
    _cell(ws, row, 8, _rnd(vaz_all / 1000), bold=True)
    _cell(ws, row, 9, _rnd(tash_all / 1000), bold=True)
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row, 1, "Jami ijtimoiy soliq").font = _BOLD_FONT
    _cell(ws, row, 7, _rnd(d["total_social"] / 1000), bold=True)

    for c_letter, w in [
        ("A", 5),
        ("B", 25),
        ("C", 12),
        ("D", 15),
        ("E", 15),
        ("F", 15),
        ("G", 15),
        ("H", 18),
        ("I", 18),
    ]:
        ws.column_dimensions[c_letter].width = w


def _build_inventar_sheet(wb: Workbook, d: dict) -> None:
    """Build the 'Inventar' sheet."""
    ws = wb.create_sheet("Inventar")
    ws.merge_cells("A1:H1")
    ws["A1"].value = "Inventar, texnika va jihozlarni xarid qilish xarajatlari"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].fill = _HEADER_FILL
    ws["A1"].alignment = _CENTER

    headers = [
        "N",
        "Mahsulotlar nomi",
        "O'lchov\nbirligi",
        "Miqdori",
        "Narxi\n(ming so'mda)",
        "Summasi\n(ming so'mda)",
    ]
    for i, h in enumerate(headers, 1):
        ws.merge_cells(start_row=3, start_column=i, end_row=4, end_column=i)
        _header_style(ws, 3, i, h)
    ws.merge_cells("G3:H3")
    _header_style(ws, 3, 7, "Moliyalashtirish manbasi")
    _header_style(ws, 4, 7, "Vazirlik hisobidan\n(ming so'mda)")
    _header_style(ws, 4, 8, "Tashkilot hisobidan\n(ming so'mda)")

    row = 5
    for idx, item in enumerate(d["inventory"], 1):
        t = float(item.price) * item.quantity
        is_v = item.financing_source == "vazirlik"
        _cell(ws, row, 1, idx)
        name_val = item.name
        if item.description:
            name_val += f"\n({item.description})"
        if item.link:
            name_val += f"\n{item.link}"
        _cell(ws, row, 2, name_val, align="left")
        _cell(ws, row, 3, item.unit)
        _cell(ws, row, 4, item.quantity)
        _cell(ws, row, 5, _rnd(float(item.price) / 1000))
        _cell(ws, row, 6, _rnd(t / 1000))
        _cell(ws, row, 7, _rnd(t / 1000) if is_v else "")
        _cell(ws, row, 8, _rnd(t / 1000) if not is_v else "")
        row += 1

    _cell(ws, row, 2, "Jami:", bold=True, align="right")
    _cell(ws, row, 6, _rnd(d["inv_total"] / 1000), bold=True)
    _cell(ws, row, 7, _rnd(d["inv_vaz"] / 1000), bold=True)
    _cell(ws, row, 8, _rnd(d["inv_tash"] / 1000), bold=True)

    for c_letter, w in [
        ("A", 5),
        ("B", 50),
        ("C", 10),
        ("D", 10),
        ("E", 12),
        ("F", 12),
        ("G", 18),
        ("H", 18),
    ]:
        ws.column_dimensions[c_letter].width = w


def _build_xom_ashyo_sheet(wb: Workbook, d: dict) -> None:
    """Build the 'Xom ashyo' (raw materials) sheet."""
    ws = wb.create_sheet("Xom ashyo")
    ws.merge_cells("A1:H1")
    ws["A1"].value = "Xomashyo va materiallarni sotib olish xarajatlari"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].fill = _HEADER_FILL
    ws["A1"].alignment = _CENTER

    headers = [
        "N",
        "Xomashyo nomi",
        "O'lchov\nbirligi",
        "Miqdori",
        "Narxi\n(ming so'mda)",
        "Summasi\n(ming so'mda)",
    ]
    for i, h in enumerate(headers, 1):
        ws.merge_cells(start_row=3, start_column=i, end_row=4, end_column=i)
        _header_style(ws, 3, i, h)
    ws.merge_cells("G3:H3")
    _header_style(ws, 3, 7, "Moliyalashtirish manbasi")
    _header_style(ws, 4, 7, "Vazirlik hisobidan\n(ming so'mda)")
    _header_style(ws, 4, 8, "Tashkilot hisobidan\n(ming so'mda)")

    row = 5
    for idx, item in enumerate(d["raw_materials"], 1):
        t = float(item.price) * item.quantity
        is_v = item.financing_source == "vazirlik"
        _cell(ws, row, 1, idx)
        _cell(ws, row, 2, item.name, align="left")
        _cell(ws, row, 3, item.unit)
        _cell(ws, row, 4, item.quantity)
        _cell(ws, row, 5, _rnd(float(item.price) / 1000))
        _cell(ws, row, 6, _rnd(t / 1000))
        _cell(ws, row, 7, _rnd(t / 1000) if is_v else "")
        _cell(ws, row, 8, _rnd(t / 1000) if not is_v else "")
        row += 1

    _cell(ws, row, 2, "Jami:", bold=True, align="right")
    _cell(ws, row, 6, _rnd(d["rm_total"] / 1000), bold=True)
    _cell(ws, row, 7, _rnd(d["rm_vaz"] / 1000), bold=True)
    _cell(ws, row, 8, _rnd(d["rm_tash"] / 1000), bold=True)

    for c_letter, w in [
        ("A", 5),
        ("B", 40),
        ("C", 10),
        ("D", 10),
        ("E", 12),
        ("F", 12),
        ("G", 18),
        ("H", 18),
    ]:
        ws.column_dimensions[c_letter].width = w


def _build_boshqa_xarajatlar_sheet(wb: Workbook, d: dict) -> None:
    """Build the 'Boshqa xar.' sheet."""
    ws = wb.create_sheet("Boshqa xar.")
    ws.merge_cells("A1:H1")
    ws["A1"].value = "Boshqa xarajatlar"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].fill = _HEADER_FILL
    ws["A1"].alignment = _CENTER

    headers = [
        "N",
        "Xarajatlar nomi",
        "O'lchov birligi",
        "Miqdori",
        "Narxi\n(ming so'mda)",
        "Summasi\n(ming so'mda)",
    ]
    for i, h in enumerate(headers, 1):
        ws.merge_cells(start_row=3, start_column=i, end_row=4, end_column=i)
        _header_style(ws, 3, i, h)
    ws.merge_cells("G3:H3")
    _header_style(ws, 3, 7, "Moliyalashtirish manbasi")
    _header_style(ws, 4, 7, "Vazirlik hisobidan\n(ming so'mda)")
    _header_style(ws, 4, 8, "Tashkilot hisobidan\n(ming so'mda)")

    row = 5
    # Management expenses
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row, 1, "Boshqa ma'muriy xarajatlar:").font = _BOLD_FONT
    _cell(ws, row, 7, _rnd(d["oe_mgmt_vaz"] / 1000))
    _cell(ws, row, 8, _rnd(d["oe_mgmt_tash"] / 1000))
    row += 1

    for idx, exp in enumerate(d["mgmt_exp"], 1):
        t = float(exp.price) * exp.quantity
        is_v = exp.financing_source == "vazirlik"
        _cell(ws, row, 1, idx)
        _cell(ws, row, 2, exp.name, align="left")
        _cell(ws, row, 3, exp.unit)
        _cell(ws, row, 4, exp.quantity)
        _cell(ws, row, 5, _rnd(float(exp.price) / 1000))
        _cell(ws, row, 6, _rnd(t / 1000))
        _cell(ws, row, 7, _rnd(t / 1000) if is_v else "")
        _cell(ws, row, 8, _rnd(t / 1000) if not is_v else "")
        row += 1

    # Production expenses
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row, 1, "Ishlab chiqarish bilan bog'liq boshqa xarajatlar:").font = (
        _BOLD_FONT
    )
    _cell(ws, row, 7, _rnd(d["oe_prod_vaz"] / 1000))
    _cell(ws, row, 8, _rnd(d["oe_prod_tash"] / 1000))
    row += 1

    for idx, exp in enumerate(d["prod_exp"], 1):
        t = float(exp.price) * exp.quantity
        is_v = exp.financing_source == "vazirlik"
        _cell(ws, row, 1, idx)
        _cell(ws, row, 2, exp.name, align="left")
        _cell(ws, row, 3, exp.unit)
        _cell(ws, row, 4, exp.quantity)
        _cell(ws, row, 5, _rnd(float(exp.price) / 1000))
        _cell(ws, row, 6, _rnd(t / 1000))
        _cell(ws, row, 7, _rnd(t / 1000) if is_v else "")
        _cell(ws, row, 8, _rnd(t / 1000) if not is_v else "")
        row += 1

    # Totals
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row, 1, "Jami:").font = _BOLD_FONT
    ws.cell(row, 1).alignment = _RIGHT
    oe_vaz = d["oe_mgmt_vaz"] + d["oe_prod_vaz"]
    oe_tash = d["oe_mgmt_tash"] + d["oe_prod_tash"]
    _cell(ws, row, 6, _rnd(d["oe_total"] / 1000), bold=True)
    _cell(ws, row, 7, _rnd(oe_vaz / 1000), bold=True)
    _cell(ws, row, 8, _rnd(oe_tash / 1000), bold=True)

    for c_letter, w in [
        ("A", 5),
        ("B", 45),
        ("C", 12),
        ("D", 10),
        ("E", 12),
        ("F", 12),
        ("G", 18),
        ("H", 18),
    ]:
        ws.column_dimensions[c_letter].width = w


def _build_tannarx_sheet(wb: Workbook, d: dict) -> None:
    """Build the 'Tannarx' (cost price) sheet."""
    ws = wb.create_sheet("Tannarx")

    ws.merge_cells("A1:C1")
    ws["A1"].value = (
        "Mahsulot (ishlar, xizmatlar)ning ishlab chiqarish tannarxiga "
        "kiritiladigan xarajatlar tarkibi."
    )
    ws["A1"].font = Font(size=10)
    ws["A1"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 60

    ws.merge_cells("A3:C3")
    ws["A3"].value = "Mahsulotlarning ishlab chiqarish tannarxi"
    ws["A3"].font = _TITLE_FONT
    ws["A3"].alignment = _CENTER

    for i, h in enumerate(["N", "Xarajatlar nomi", "Summasi\n(ming so'mda)"], 1):
        _header_style(ws, 5, i, h)

    prod_social = d["prod_total"] * float(SOCIAL_TAX_RATE)
    items = [
        ("Ishlab chiqarish xodimlarining ish haqi", d["prod_total"]),
        ("Ijtimoiy soliq", prod_social),
        (
            "Xom ashyo va materiallarni sotib olish bilan " "bog'liq xarajatlar",
            d["rm_total"],
        ),
        ("Asosiy vositalarning amortizatsiya xarajatlari", d["amortizatsiya"]),
        ("Boshqa ishlab chiqarish xarajatlari", d["oe_prod_total"]),
    ]

    row = 6
    for idx, (name, val) in enumerate(items, 1):
        _cell(ws, row, 1, idx)
        _cell(ws, row, 2, name, align="left")
        _cell(ws, row, 3, _rnd(val / 1000))
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row, 1, "Jami tannarx").font = _BOLD_FONT
    ws.cell(row, 1).alignment = _RIGHT
    _cell(ws, row, 3, _rnd(d["jami_tannarx"] / 1000), bold=True)
    row += 2

    for idx, product in enumerate(d["products"]):
        tpq = d["total_products_qty"]
        unit_cost = (
            (d["jami_tannarx"] * (product.quantity / tpq)) / product.quantity
            if tpq > 0
            else 0
        )
        total_cost = d["jami_tannarx"] * (product.quantity / tpq) if tpq > 0 else 0
        ws.cell(row, 2, f"{product.name}*").alignment = _LEFT
        ws.cell(row, 3, _rnd(total_cost / 1000))
        row += 1
        ws.cell(row, 1, f"{idx + 1}-mahsulot").font = _BOLD_FONT
        ws.cell(row, 2, "Mahsulot soni")
        ws.cell(row, 3, product.quantity)
        row += 1
        ws.cell(row, 2, "Mahsulot narxi").font = _BOLD_FONT
        ws.cell(row, 3, _rnd(unit_cost / 1000)).font = _BOLD_FONT
        row += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 18


def _build_davr_xarajatlari_sheet(wb: Workbook, d: dict) -> None:
    """Build the 'Davr xarajatlari' sheet."""
    ws = wb.create_sheet("Davr xarajatlari")
    ws.merge_cells("A1:C1")
    ws["A1"].value = "DAVR XARAJATLARI"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = _CENTER

    ws.merge_cells("A2:C2")
    ws["A2"].value = "(ming so'mda)"
    ws["A2"].alignment = _CENTER

    for i, h in enumerate(["N", "Xarajatlar nomi", "Summasi"], 1):
        _header_style(ws, 4, i, h)

    row = 5
    for idx, exp in enumerate(d["davr"], 1):
        _cell(ws, row, 1, idx)
        _cell(ws, row, 2, exp.name, align="left")
        _cell(ws, row, 3, _rnd(float(exp.amount)))
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row, 1, "JAMI DAVR XARAJATLARI").font = _BOLD_FONT
    ws.cell(row, 1).alignment = _RIGHT
    _cell(ws, row, 3, _rnd(d["davr_total"]), bold=True)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 18


def _build_sotish_rejasi_sheet(wb: Workbook, d: dict) -> None:
    """Build the 'Sotish rejasi' sheet."""
    ws = wb.create_sheet("Sotish rejasi")
    ws.merge_cells("A1:F1")
    ws["A1"].value = "SOTISH REJASI"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = _CENTER

    ws.merge_cells("A2:F2")
    ws["A2"].value = "(ming so'mda)"
    ws["A2"].alignment = _CENTER

    row = 4
    grand_rev = 0.0

    for year_num in range(1, d["project_years"] + 1):
        year_obj = next(
            (y for y in d["sotish_yillari"] if y.year == year_num),
            None,
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row, 1, f"{year_num}-YIL")
        c.font = Font(bold=True, size=12)
        c.alignment = _CENTER
        c.fill = PatternFill("solid", fgColor="E8E8E8")
        row += 1

        for i, h in enumerate(
            ["N", "Mahsulotlar nomi", "O'lch. birligi", "Miqdori", "Narxi", "Summasi"],
            1,
        ):
            _header_style(ws, row, i, h)
        row += 1

        year_revenue = 0.0
        if year_obj:
            prods = list(year_obj.products.all())
            for idx, p in enumerate(prods, 1):
                rev = p.quantity * float(p.price)
                _cell(ws, row, 1, idx)
                _cell(ws, row, 2, p.name, align="left")
                _cell(ws, row, 3, p.unit)
                _cell(ws, row, 4, p.quantity)
                _cell(ws, row, 5, _rnd(float(p.price)))
                _cell(ws, row, 6, _rnd(rev))
                year_revenue += rev
                row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row, 1, f"JAMI {year_num}-YIL").font = _BOLD_FONT
        ws.cell(row, 1).alignment = _RIGHT
        _cell(ws, row, 6, _rnd(year_revenue), bold=True)
        grand_rev += year_revenue
        row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row, 1, "UMUMIY JAMI").font = Font(bold=True, size=12)
    ws.cell(row, 1).alignment = _RIGHT
    ws.cell(row, 1).fill = _HEADER_FILL
    c = ws.cell(row, 6, _rnd(grand_rev))
    c.font = Font(bold=True, size=12)
    c.fill = _HEADER_FILL
    c.border = _BORDER

    for c_letter, w in [
        ("A", 6),
        ("B", 35),
        ("C", 12),
        ("D", 12),
        ("E", 15),
        ("F", 18),
    ]:
        ws.column_dimensions[c_letter].width = w


def _build_moliyaviy_xisobot_sheet(wb: Workbook, d: dict) -> None:
    """Build the 'Moliyaviy xisobot' sheet."""
    ws = wb.create_sheet("Moliyaviy xisobot")
    py = d["project_years"]
    last_col = py + 2

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws["A1"].value = "MOLIYAVIY XISOBOT (FOYDA-ZARAR)"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = _CENTER

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws["A2"].value = "(ming so'mda)"
    ws["A2"].alignment = _CENTER

    _header_style(ws, 4, 1, "N")
    _header_style(ws, 4, 2, "Ko'rsatkichlar")
    for i in range(1, py + 1):
        _header_style(ws, 4, i + 2, f"{i}-yil")

    # Prepare yearly data
    yearly_tannarx = d["jami_tannarx"] / py if py else 0
    yearly_davr = d["davr_total"] / py if py else 0
    st = float(PROFIT_TAX_RATE)

    sotish_by_year: dict[int, float] = {}
    for yil in d["sotish_yillari"]:
        rev = sum(p.quantity * float(p.price) for p in yil.products.all())
        sotish_by_year[yil.year] = rev

    indicators = [
        "Sotishdan tushum",
        "Sotilgan mahsulot tannarxi",
        "Yalpi daromad (foyda)",
        "Davr xarajatlari",
        "Asosiy faoliyat foydasi",
        "Foyda solig'i (12%)",
        "SOF FOYDA",
    ]

    total_profit = 0.0
    row = 5
    for idx, name in enumerate(indicators, 1):
        is_bold = name in (
            "Yalpi daromad (foyda)",
            "Asosiy faoliyat foydasi",
            "SOF FOYDA",
        )
        _cell(ws, row, 1, idx)
        _cell(ws, row, 2, name, bold=is_bold, align="left")
        for yr in range(1, py + 1):
            sotish = sotish_by_year.get(yr, 0)
            yalpi = sotish - yearly_tannarx
            asosiy = yalpi - yearly_davr
            foyda_soliq = asosiy * st if asosiy > 0 else 0
            sof = asosiy - foyda_soliq

            val_map = {
                0: sotish,
                1: yearly_tannarx,
                2: yalpi,
                3: yearly_davr,
                4: asosiy,
                5: foyda_soliq,
                6: sof,
            }
            val = val_map[idx - 1]
            c = ws.cell(row, yr + 2, _rnd(val / 1000))
            c.border = _BORDER
            c.alignment = _CENTER
            if name == "SOF FOYDA":
                c.font = Font(bold=True, size=11)
                c.fill = PatternFill(
                    "solid",
                    fgColor="90EE90" if val >= 0 else "FF9999",
                )
                if yr == py:
                    total_profit += sof
            elif is_bold:
                c.font = _BOLD_FONT
            if idx == 7:
                total_profit = total_profit + sof if yr < py else total_profit
        row += 1

    # Recalculate total profit properly
    total_profit = 0.0
    for yr in range(1, py + 1):
        sotish = sotish_by_year.get(yr, 0)
        yalpi = sotish - yearly_tannarx
        asosiy = yalpi - yearly_davr
        foyda_soliq = asosiy * st if asosiy > 0 else 0
        total_profit += asosiy - foyda_soliq

    row += 1
    ws.cell(row, 2, "Umumiy sof foyda (barcha yillar)").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=last_col)
    c = ws.cell(row, 3, _rnd(total_profit / 1000))
    c.font = Font(bold=True, size=12)
    c.alignment = _CENTER
    c.fill = PatternFill(
        "solid",
        fgColor="90EE90" if total_profit >= 0 else "FF9999",
    )
    c.border = _BORDER

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 35
    for i in range(1, py + 1):
        ws.column_dimensions[get_column_letter(i + 2)].width = 15


# ------------------------------------------------------------------
# Public API
# ------------------------------------------------------------------
def generate_smeta_excel(smeta: XarajatlarSmetasi) -> str:
    """Generate an Excel workbook for the given smeta and save it.

    Args:
        smeta: The XarajatlarSmetasi instance.

    Returns:
        The relative file path of the saved Excel file.
    """
    d = _gather_smeta_data(smeta)

    wb = Workbook()
    _build_jami_sheet(wb, d)
    _build_ish_haqi_sheet(wb, d)
    _build_inventar_sheet(wb, d)
    _build_xom_ashyo_sheet(wb, d)
    _build_boshqa_xarajatlar_sheet(wb, d)
    _build_tannarx_sheet(wb, d)
    _build_davr_xarajatlari_sheet(wb, d)
    _build_sotish_rejasi_sheet(wb, d)
    _build_moliyaviy_xisobot_sheet(wb, d)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = (
        smeta.project_name.replace("/", "_").replace("\\", "_").replace(":", "_")[:120]
        or "Xarajatlar_smetasi"
    )
    filename = f"{safe_name}.xlsx"

    smeta.excel_file.save(filename, ContentFile(buf.read()), save=True)
    logger.info("Excel fayl yaratildi: smeta_id=%d", smeta.pk)
    return smeta.excel_file.url
